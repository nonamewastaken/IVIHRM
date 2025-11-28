from flask import request, jsonify, render_template, redirect, session
from models import User, Attendance, db
from datetime import datetime, date, timedelta
from core.auth import login_required, api_login_required
from . import attendance_bp
import openpyxl
import os
import tempfile
import calendar
import re
import json
from werkzeug.utils import secure_filename

@attendance_bp.route('/')
@login_required
def attendance_dashboard():
    """Main attendance dashboard page - redirects to overview"""
    return redirect('/attendance/overview')

@attendance_bp.route('/overview')
@login_required
def overview():
    """Attendance Overview page"""
    user = User.query.get(session['user_id'])
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('overview.html', user=user)

@attendance_bp.route('/admin')
@login_required
def attendance_admin():
    """Admin attendance management page - redirects to monthly detail"""
    user = User.query.get(session['user_id'])
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    # TODO: Add admin role check here
    # if not user.is_admin:
    #     return redirect('/attendance')
    
    # Redirect to monthly attendance detail for now
    return redirect('/attendance/monthly_attendance_detail')

@attendance_bp.route('/monthly_attendance_detail')
@login_required
def monthly_attendance_detail():
    """Monthly Attendance Detail page"""
    user = User.query.get(session['user_id'])
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('monthly_attendance_detail.html', user=user)

@attendance_bp.route('/attendance_summary')
@login_required
def attendance_summary():
    """Attendance Summary page"""
    user = User.query.get(session['user_id'])
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('attendance_summary.html', user=user)

@attendance_bp.route('/work_data')
@login_required
def work_data():
    """Work Data page"""
    user = User.query.get(session['user_id'])
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('work_data.html', user=user)

@attendance_bp.route('/history')
@login_required
def history():
    """Attendance History page"""
    user = User.query.get(session['user_id'])
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('attendance_history.html', user=user)

@attendance_bp.route('/api/check-in', methods=['POST'])
@api_login_required
def check_in():
    """Employee check-in API"""
    try:
        user = User.query.get(session['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        current_time = datetime.now()
        today = current_time.date()
        
        # Check if already checked in today
        existing_checkin = Attendance.query.filter(
            Attendance.user_id == user.id,
            Attendance.date == today,
            Attendance.check_in_time.isnot(None)
        ).first()
        
        if existing_checkin:
            return jsonify({'error': 'Already checked in today'}), 400
        
        # Create new attendance record
        attendance = Attendance(
            user_id=user.id,
            date=today,
            check_in_time=current_time,
            status='present'
        )
        
        db.session.add(attendance)
        db.session.commit()
        
        return jsonify({
            'message': 'Check-in successful',
            'check_in_time': current_time.strftime('%H:%M:%S'),
            'date': today.strftime('%Y-%m-%d')
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Check-in failed'}), 500

@attendance_bp.route('/api/check-out', methods=['POST'])
@api_login_required
def check_out():
    """Employee check-out API"""
    try:
        user = User.query.get(session['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        current_time = datetime.now()
        today = current_time.date()
        
        # Find today's attendance record
        attendance = Attendance.query.filter(
            Attendance.user_id == user.id,
            Attendance.date == today,
            Attendance.check_in_time.isnot(None)
        ).first()
        
        if not attendance:
            return jsonify({'error': 'No check-in found for today'}), 400
        
        if attendance.check_out_time:
            return jsonify({'error': 'Already checked out today'}), 400
        
        # Update attendance record
        attendance.check_out_time = current_time
        attendance.work_hours = calculate_work_hours(attendance.check_in_time, current_time)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Check-out successful',
            'check_out_time': current_time.strftime('%H:%M:%S'),
            'work_hours': attendance.work_hours
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Check-out failed'}), 500

@attendance_bp.route('/api/attendance-status', methods=['GET'])
@api_login_required
def attendance_status():
    """Get current attendance status for today"""
    try:
        user = User.query.get(session['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        today = date.today()
        attendance = Attendance.query.filter_by(
            user_id=user.id,
            date=today
        ).first()
        
        if not attendance:
            return jsonify({
                'status': 'not_checked_in',
                'message': 'Not checked in today'
            }), 200
        
        if attendance.check_in_time and not attendance.check_out_time:
            return jsonify({
                'status': 'checked_in',
                'check_in_time': attendance.check_in_time.strftime('%H:%M:%S'),
                'message': 'Checked in, ready to check out'
            }), 200
        
        if attendance.check_in_time and attendance.check_out_time:
            return jsonify({
                'status': 'checked_out',
                'check_in_time': attendance.check_in_time.strftime('%H:%M:%S'),
                'check_out_time': attendance.check_out_time.strftime('%H:%M:%S'),
                'work_hours': attendance.work_hours,
                'message': 'Completed for today'
            }), 200
        
    except Exception as e:
        return jsonify({'error': 'Failed to get attendance status'}), 500

@attendance_bp.route('/api/attendance-history', methods=['GET'])
@api_login_required
def attendance_history():
    """Get attendance history for the current user"""
    try:
        user = User.query.get(session['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Get date range from query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            # Default to last 30 days
            end_date = date.today()
            start_date = end_date - timedelta(days=30)
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Query attendance records
        attendance_records = Attendance.query.filter(
            Attendance.user_id == user.id,
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).order_by(Attendance.date.desc()).all()
        
        # Format response
        history = []
        for record in attendance_records:
            history.append({
                'date': record.date.strftime('%Y-%m-%d'),
                'check_in_time': record.check_in_time.strftime('%H:%M:%S') if record.check_in_time else None,
                'check_out_time': record.check_out_time.strftime('%H:%M:%S') if record.check_out_time else None,
                'work_hours': record.work_hours,
                'status': record.status,
                'notes': record.notes
            })
        
        return jsonify({
            'history': history,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d')
        }), 200
        
    except Exception as e:
        return jsonify({'error': 'Failed to get attendance history'}), 500

@attendance_bp.route('/attendance-history')
@login_required
def attendance_history_page():
    """Attendance history page"""
    user = User.query.get(session['user_id'])
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('attendance_history.html', user=user)

def calculate_work_hours(check_in_time, check_out_time):
    """Calculate work hours between check-in and check-out"""
    if not check_in_time or not check_out_time:
        return 0
    
    time_diff = check_out_time - check_in_time
    hours = time_diff.total_seconds() / 3600
    return round(hours, 2)


# ========================= Timesheet Boards (Summary) =========================
@attendance_bp.route('/api/timesheet-boards', methods=['GET'])
@api_login_required
def list_timesheet_boards():
    """List timesheet summary boards created by managers.

    This is a stub endpoint that currently returns an empty list with
    pagination metadata and echoes back filter parameters. It is intended to
    power the Summary tab UI. Later, this can be wired to real persistence.
    """
    try:
        # Parse filters
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 10))
        query_text = request.args.get('q')
        unit = request.args.get('unit')
        board_type = request.args.get('type')
        status = request.args.get('status')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')

        # Placeholder data (empty state by default)
        items = []
        total = 0

        # Response structure
        return jsonify({
            'items': items,
            'page': page,
            'page_size': page_size,
            'total': total,
            'filters': {
                'q': query_text,
                'unit': unit,
                'type': board_type,
                'status': status,
                'from_date': from_date,
                'to_date': to_date
            }
        }), 200
    except Exception:
        return jsonify({'error': 'Failed to load timesheet boards'}), 500


@attendance_bp.route('/api/work-data', methods=['GET'])
@api_login_required
def list_work_data():
    """List raw attendance work data records per employee (stub).

    Accepts filters similar to the example UI and returns a paginated list.
    """
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 10))
        q = request.args.get('q')
        department = request.args.get('department')
        method = request.args.get('method')
        shift = request.args.get('shift')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')

        items = []
        total = 0

        return jsonify({
            'items': items,
            'page': page,
            'page_size': page_size,
            'total': total,
            'filters': {
                'q': q,
                'department': department,
                'method': method,
                'shift': shift,
                'from_date': from_date,
                'to_date': to_date
            }
        }), 200
    except Exception:
        return jsonify({'error': 'Failed to load work data'}), 500

@attendance_bp.route('/api/monthly-data', methods=['GET'])
@api_login_required
def get_monthly_data():
    """Get monthly attendance data for the selected month"""
    try:
        month_year = request.args.get('month_year')
        if not month_year:
            return jsonify({'error': 'Month and year are required'}), 400
        
        try:
            year, month = month_year.split('-')
            year, month = int(year), int(month)
        except ValueError:
            return jsonify({'error': 'Invalid month/year format. Use YYYY-MM'}), 400
        
        # Calculate days in month
        import calendar
        days_in_month = calendar.monthrange(year, month)[1]
        
        # Query attendance records for the specified month
        attendance_records = Attendance.query.filter(
            Attendance.year == year,
            Attendance.month == month
        ).order_by(Attendance.employee_no).all()
        
        # Format employee data
        employees = []
        for i, record in enumerate(attendance_records):
            employee_data = {
                'employee_no': record.employee_no,
                'employee_name': record.employee_name,
                'year': record.year,
                'month': record.month,
                'daily_attendance': record.daily_attendance,
                'other_data': record.other_data,
                'created_at': record.created_at.isoformat() if record.created_at else None
            }
            employees.append(employee_data)
            
            # Debug first few records
            if i < 2:  # Only debug first 2 records
                print(f"DEBUG API: Employee {i+1} - {record.employee_no} - {record.employee_name}")
                print(f"  Daily attendance: {record.daily_attendance[:100]}...")
                print(f"  Other data length: {len(record.other_data) if record.other_data else 0}")
                print(f"  Other data sample: {record.other_data[:200] if record.other_data else 'None'}...")
        
        return jsonify({
            'success': True,
            'employees': employees,
            'daysInMonth': days_in_month,
            'month': f"{year}-{month:02d}",
            'total_employees': len(employees)
        }), 200
        
    except Exception as e:
        print(f"Error getting monthly data: {e}")
        return jsonify({'error': 'Failed to load monthly data'}), 500


@attendance_bp.route('/api/import-excel', methods=['POST'])
@api_login_required
def import_excel():
    """Import attendance data from Excel file with comprehensive validation"""
    try:
        print("DEBUG: Import Excel function called")  # Debug log
        # 1. UPFRONT VALIDATION
        
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # MIME type validation (.xlsx only)
        if not file.filename.lower().endswith('.xlsx'):
            return jsonify({'error': 'Invalid file format. Only .xlsx files are allowed'}), 400
        
        # File size validation (10MB limit)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        if file_size > 10 * 1024 * 1024:  # 10MB
            return jsonify({'error': 'File too large. Maximum size is 10MB'}), 400
        
        # Get month and year from request
        month_year = request.form.get('month_year', '')
        if not month_year:
            return jsonify({'error': 'Month and year are required'}), 400
        
        try:
            year, month = month_year.split('-')
            year, month = int(year), int(month)
        except ValueError:
            return jsonify({'error': 'Invalid month/year format. Use YYYY-MM'}), 400
        
        # Save file temporarily
        filename = secure_filename(file.filename)
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        
        try:
            # 2. PREPROCESSING AND VALIDATION
            
            # Open Excel file
            workbook = openpyxl.load_workbook(temp_file.name, data_only=True)
            worksheet = workbook.active
            
            # Log Excel file information
            print(f"📊 EXCEL FILE ANALYSIS:")
            print(f"   File: {file.filename}")
            print(f"   Size: {file_size:,} bytes")
            print(f"   Total Rows: {worksheet.max_row}")
            print(f"   Total Columns: {worksheet.max_column}")
            print(f"   Target Month: {year}-{month:02d}")
            
            # Validate Excel format with comprehensive checks
            validation_result = validate_excel_format_comprehensive(worksheet, year, month)
            if not validation_result['valid']:
                error_message = 'Invalid Excel format:\n'
                for i, error in enumerate(validation_result['errors'], 1):
                    error_message += f"{i}. {error}\n"
                
                return jsonify({
                    'error': error_message.strip(),
                    'details': validation_result['errors'],
                    'validation_info': {
                        'expected_columns': validation_result.get('expected_columns', 0),
                        'actual_columns': worksheet.max_column,
                        'expected_rows': 29,  # 4 headers + 25 data
                        'actual_rows': worksheet.max_row,
                        'days_in_month': validation_result.get('days_in_month', 0),
                        'data_rows_found': validation_result.get('data_rows_found', 0)
                    }
                }), 400
            
            # 3. PROCESS DATA
            
            print(f"🔍 DATA PROCESSING:")
            print(f"   Expected columns: {validation_result.get('expected_columns', 0)}")
            print(f"   Days in month: {validation_result.get('days_in_month', 0)}")
            print(f"   Data rows found: {validation_result.get('data_rows_found', 0)}")
            
            # Process Excel data with security measures
            processed_data = process_excel_data_secure(worksheet, year, month)
            
            print(f"✅ PROCESSING COMPLETE:")
            print(f"   Employees processed: {len(processed_data)}")
            if processed_data:
                sample_employee = processed_data[0]
                daily_attendance_count = len(sample_employee.get('daily_attendance', []))
                other_data_count = len(sample_employee.get('other_data', {}))
                print(f"   📊 COLUMN BREAKDOWN PER EMPLOYEE:")
                print(f"     - STT (Row number): 1 column")
                print(f"     - Employee Name: 1 column")
                print(f"     - Daily attendance: {daily_attendance_count} columns")
                print(f"     - Other data: {other_data_count} columns")
                print(f"     - TOTAL COLUMNS PER EMPLOYEE: {2 + daily_attendance_count + other_data_count}")
                print(f"   Sample employee: {sample_employee.get('employee_no', 'N/A')} - {sample_employee.get('employee_name', 'N/A')}")
            
            # 4. DATABASE OPERATIONS
            
            print(f"💾 DATABASE OPERATIONS:")
            print(f"   Clearing existing data for {year}-{month:02d}...")
            
            # Clear existing data for the month (replace mode)
            clear_monthly_data(year, month)
            
            print(f"   Saving {len(processed_data)} employee records to database...")
            
            # Save processed data to database
            save_attendance_data(processed_data, year, month)
            
            print(f"🎉 IMPORT SUCCESSFUL!")
            print(f"   Total employees imported: {len(processed_data)}")
            print(f"   Month: {year}-{month:02d}")
            print(f"   Database records created: {len(processed_data)}")
            
            return jsonify({
                'success': True,
                'message': 'Excel file imported successfully',
                'rows_imported': len(processed_data),
                'month': f"{year}-{month:02d}",
                'details': {
                    'employees': len(processed_data),
                    'days_in_month': validation_result.get('days_in_month', 0)
                }
            }), 200
            
        finally:
            # Clean up temporary file
            os.unlink(temp_file.name)
            
    except Exception as e:
        print(f"DEBUG: Import error: {str(e)}")  # Debug log
        import traceback
        traceback.print_exc()  # Print full traceback
        return jsonify({'error': f'Import failed: {str(e)}'}), 500


def validate_excel_format_comprehensive(worksheet, year, month):
    """Comprehensive validation of Excel file format according to requirements"""
    errors = []
    
    # Calculate expected dimensions
    days_in_month = calendar.monthrange(year, month)[1]
    expected_columns = 29 + days_in_month  # 29 fixed columns + days in month
    max_data_rows = 25  # Maximum 25 employee rows
    
    # 1. Check minimum rows (4 headers + at least 1 data row)
    if worksheet.max_row < 5:  # 4 headers + at least 1 data row
        errors.append(f"Excel file must have at least 5 rows (4 headers + at least 1 data row). Found: {worksheet.max_row}")
        return {'valid': False, 'errors': errors, 'days_in_month': days_in_month, 'expected_columns': expected_columns, 'data_rows_found': 0}
    
    # 2. Check column count
    if worksheet.max_column < expected_columns:
        errors.append(f"Expected {expected_columns} columns (29 fixed + {days_in_month} days). Found: {worksheet.max_column}")
        return {'valid': False, 'errors': errors, 'days_in_month': days_in_month, 'expected_columns': expected_columns, 'data_rows_found': 0}
    
    # 3. Skip header validation - we ignore the first 4 rows completely
    
    # 4. Check for merged cells in data area (rows 5 onwards, skipping first 4 header rows)
    merged_cells_found = []
    max_data_row = min(5 + max_data_rows, worksheet.max_row + 1)  # Up to 25 data rows after headers
    for row in range(5, max_data_row):  # Data rows
        for col in range(1, min(expected_columns + 1, worksheet.max_column + 1)):
            cell = worksheet.cell(row=row, column=col)
            if hasattr(cell, 'merged_cell') and cell.merged_cell:
                merged_cells_found.append(f"Row {row}, Column {col}")
    
    if merged_cells_found:
        errors.append(f"Merged cells found in data area: {', '.join(merged_cells_found[:5])}{'...' if len(merged_cells_found) > 5 else ''}. Data area must not contain merged cells.")
    
    # 5. Check for formulas in data area (rows 5 onwards, skipping first 4 header rows)
    formulas_found = []
    for row in range(5, max_data_row):  # Data rows
        for col in range(1, min(expected_columns + 1, worksheet.max_column + 1)):
            cell = worksheet.cell(row=row, column=col)
            if cell.data_type == 'f':  # Formula
                formulas_found.append(f"Row {row}, Column {col}")
    
    if formulas_found:
        errors.append(f"Formulas found in data area: {', '.join(formulas_found[:5])}{'...' if len(formulas_found) > 5 else ''}. Data area must not contain formulas.")
    
    # 6. Validate data rows count (1 to 25 employees, skipping first 4 header rows)
    data_rows = 0
    empty_rows = []
    for row in range(5, max_data_row):  # Check data rows
        # Check if row has any data
        has_data = any(worksheet.cell(row=row, column=col).value is not None for col in range(1, min(expected_columns + 1, worksheet.max_column + 1)))
        if has_data:
            data_rows += 1
        else:
            empty_rows.append(row)
    
    if data_rows == 0:
        errors.append(f"No employee data found. Expected at least 1 employee row.")
    elif data_rows > max_data_rows:
        errors.append(f"Too many employees. Maximum {max_data_rows} employees allowed. Found: {data_rows}")
    elif empty_rows and data_rows < max_data_rows:
        # Only show empty rows warning if we haven't reached the maximum
        errors.append(f"Empty rows found: {empty_rows[:10]}{'...' if len(empty_rows) > 10 else ''}. Consider filling them with employee data.")
    
    # 7. Check for required data in first data row (row 5, skipping first 4 header rows)
    if data_rows > 0:
        # Check if first data row has employee number and name
        first_data_row = 5
        employee_no = worksheet.cell(row=first_data_row, column=1).value
        employee_name = worksheet.cell(row=first_data_row, column=2).value
        
        if not employee_no:
            errors.append(f"First data row (row {first_data_row}) is missing employee number in column 1")
        if not employee_name:
            errors.append(f"First data row (row {first_data_row}) is missing employee name in column 2")
    
    return {
        'valid': len(errors) == 0, 
        'errors': errors, 
        'days_in_month': days_in_month,
        'expected_columns': expected_columns,
        'data_rows_found': data_rows
    }


def process_excel_data(worksheet, month):
    """Process Excel data and return structured attendance data"""
    processed_data = []
    
    # Get number of days in the month
    year = datetime.now().year
    if month in ['1', '3', '5', '7', '8', '10', '12']:
        days_in_month = 31
    elif month in ['4', '6', '9', '11']:
        days_in_month = 30
    else:  # February
        days_in_month = 29 if year % 4 == 0 else 28
    
    # Process data rows (starting from row 5, after 4 header rows)
    for row_num in range(5, min(worksheet.max_row + 1, 30)):  # Max 25 data rows
        row_data = {}
        
        # Get basic info
        row_data['row_number'] = worksheet.cell(row=row_num, column=1).value
        row_data['employee_name'] = worksheet.cell(row=row_num, column=2).value
        
        if not row_data['employee_name']:
            continue  # Skip empty rows
        
        # Get daily attendance data (columns 3 to 3+days_in_month-1)
        daily_data = []
        for day in range(1, days_in_month + 1):
            cell_value = worksheet.cell(row=row_num, column=2 + day).value
            daily_data.append(cell_value if cell_value is not None else '')
        
        row_data['daily_attendance'] = daily_data
        
        # Get other columns data
        col_offset = 2 + days_in_month
        row_data['standard_workdays'] = worksheet.cell(row=row_num, column=col_offset).value
        row_data['admin_workdays'] = worksheet.cell(row=row_num, column=col_offset + 1).value
        row_data['business_trip_workdays'] = worksheet.cell(row=row_num, column=col_offset + 2).value
        
        # Overtime data (4 columns)
        row_data['overtime_o1'] = worksheet.cell(row=row_num, column=col_offset + 3).value
        row_data['overtime_o2'] = worksheet.cell(row=row_num, column=col_offset + 4).value
        row_data['overtime_o3'] = worksheet.cell(row=row_num, column=col_offset + 5).value
        row_data['overtime_total'] = worksheet.cell(row=row_num, column=col_offset + 6).value
        
        # Continue with other columns...
        # (This is a simplified version - you can expand based on your exact needs)
        
        processed_data.append(row_data)
    
    return processed_data


def process_excel_data_secure(worksheet, year, month):
    """Process Excel data with security measures and return structured attendance data"""
    import re
    processed_data = []
    
    # Calculate days in month
    days_in_month = calendar.monthrange(year, month)[1]
    expected_columns = 29 + days_in_month
    
    print(f"📋 COLUMN ANALYSIS:")
    print(f"   Days in month: {days_in_month}")
    print(f"   Expected total columns: {expected_columns}")
    print(f"   Actual worksheet columns: {worksheet.max_column}")
    print(f"   Column breakdown:")
    print(f"     - STT (Row number): 1 column (column 1)")
    print(f"     - Employee Name: 1 column (column 2)")
    print(f"     - Daily attendance columns: {days_in_month} (columns 3-{2+days_in_month})")
    print(f"     - Other data columns: 29 (columns {3+days_in_month}-{expected_columns})")
    print(f"     - TOTAL EXPECTED: {2 + days_in_month + 29} columns per employee")
    print(f"     - ACTUAL WORKSHEET: {worksheet.max_column} columns")
    
    # Process data rows (skip first 4 header rows, up to 25 employees)
    max_data_row = min(5 + 25, worksheet.max_row + 1)  # Up to 25 data rows after headers
    for row in range(5, max_data_row):  # Up to 25 data rows
        row_data = {}
        
        # Employee number (column 1)
        employee_no = worksheet.cell(row=row, column=1).value
        if employee_no:
            # Security: strip leading formula characters
            employee_no_str = str(employee_no).strip()
            employee_no_str = re.sub(r'^[=+\-@]', '', employee_no_str)
            row_data['employee_no'] = employee_no_str
        else:
            # Use row number as employee number if column 1 is empty
            row_data['employee_no'] = f"EMP_{row-4:03d}"  # EMP_001, EMP_002, etc.
            print(f"DEBUG: Row {row} has no employee number, using {row_data['employee_no']}")
        
        # Employee name (column 2)
        employee_name = worksheet.cell(row=row, column=2).value
        if employee_name:
            # Security: strip leading formula characters
            employee_name_str = str(employee_name).strip()
            employee_name_str = re.sub(r'^[=+\-@]', '', employee_name_str)
            row_data['employee_name'] = employee_name_str
        else:
            row_data['employee_name'] = ''
        
        # Process daily attendance (columns 3 to 2+days_in_month)
        daily_attendance = []
        for day in range(1, days_in_month + 1):
            col = 2 + day  # Start from column 3
            if col <= expected_columns:
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None:
                    # Security: strip leading formula characters
                    cell_str = str(cell_value).strip()
                    cell_str = re.sub(r'^[=+\-@]', '', cell_str)
                    daily_attendance.append(cell_str)
                else:
                    daily_attendance.append('')
            else:
                daily_attendance.append('')
        
        row_data['daily_attendance'] = daily_attendance
        
        # Process other columns (overtime, leave, etc.) - columns after daily attendance
        other_data = {}
        other_data_start_col = 2 + days_in_month + 1  # Start after daily attendance columns
        for col in range(other_data_start_col, expected_columns + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None:
                # Security: strip leading formula characters
                cell_str = str(cell_value).strip()
                cell_str = re.sub(r'^[=+\-@]', '', cell_str)
                other_data[f'col_{col}'] = cell_str
            else:
                other_data[f'col_{col}'] = ''
        
        row_data['other_data'] = other_data
        row_data['year'] = year
        row_data['month'] = month
        
        # Log column data for this employee
        if row == 5:  # Only log for first employee to avoid spam
            print(f"   📊 SAMPLE EMPLOYEE DATA ({employee_no_str} - {employee_name_str}):")
            print(f"     - STT: 1 column")
            print(f"     - Employee Name: 1 column")
            print(f"     - Daily attendance: {len(daily_attendance)} columns (columns 3-{2+days_in_month})")
            print(f"     - Other data: {len(other_data)} columns (columns {other_data_start_col}-{expected_columns})")
            print(f"     - TOTAL COLUMNS: {2 + len(daily_attendance) + len(other_data)}")
            print(f"     - Daily attendance sample: {daily_attendance[:5]}...")
            print(f"     - Other data sample: {dict(list(other_data.items())[:5])}...")
        
        processed_data.append(row_data)
    
    return processed_data


def clear_monthly_data(year, month):
    """Clear existing attendance data for the specified month"""
    try:
        from sqlalchemy import inspect
        # Check if the table exists before trying to delete
        if inspect(db.engine).has_table('attendance'):
            # Clear data for the specific month/year
            Attendance.query.filter(
                Attendance.year == year,
                Attendance.month == month
            ).delete()
            db.session.commit()
            print(f"Cleared existing data for {year}-{month:02d}")
        else:
            print("Attendance table doesn't exist yet, skipping clear operation")
    except Exception as e:
        print(f"Error clearing monthly data: {e}")
        db.session.rollback()
        raise e


def save_attendance_data(processed_data, year, month):
    """Save processed attendance data to database with batch operations"""
    try:
        attendance_records = []
        
        print(f"DEBUG: Saving {len(processed_data)} attendance records to database")
        
        for i, data in enumerate(processed_data):
            # Create one record per employee with all their data
            attendance = Attendance(
                employee_no=data.get('employee_no', ''),
                employee_name=data.get('employee_name', ''),
                year=year,
                month=month,
                daily_attendance=','.join(data.get('daily_attendance', [])),
                other_data=json.dumps(data.get('other_data', {})),
                created_at=datetime.now()
            )
            attendance_records.append(attendance)
            
            # Debug first few records in detail
            if i < 3:  # Only debug first 3 records to avoid spam
                print(f"DEBUG: Record {i+1} - {data.get('employee_no')} - {data.get('employee_name')}")
                print(f"  Daily attendance length: {len(data.get('daily_attendance', []))}")
                print(f"  Other data keys: {list(data.get('other_data', {}).keys())}")
                print(f"  Other data count: {len(data.get('other_data', {}))}")
                print(f"  Sample other data: {dict(list(data.get('other_data', {}).items())[:5])}")
        
        # Batch insert for performance
        db.session.add_all(attendance_records)
        db.session.commit()
        
        print(f"DEBUG: Successfully saved {len(attendance_records)} records to database")
        
        # Verify data was saved by querying it back
        saved_records = Attendance.query.filter(
            Attendance.year == year,
            Attendance.month == month
        ).all()
        
        print(f"DEBUG: Verification - Found {len(saved_records)} records in database for {year}-{month:02d}")
        if saved_records:
            sample_record = saved_records[0]
            print(f"DEBUG: Sample saved record:")
            print(f"  Employee: {sample_record.employee_no} - {sample_record.employee_name}")
            print(f"  Daily attendance length: {len(sample_record.daily_attendance.split(',')) if sample_record.daily_attendance else 0}")
            print(f"  Other data length: {len(sample_record.other_data) if sample_record.other_data else 0}")
            print(f"  Other data sample: {sample_record.other_data[:200]}...")
        
    except Exception as e:
        print(f"DEBUG: Error saving attendance data: {e}")
        db.session.rollback()
        raise e

@attendance_bp.route('/checkin')
@login_required
def attendance_checkin():
    user = User.query.get(session['user_id'])
    
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('under_development.html', user=user, page_title="Check In")

@attendance_bp.route('/timeclock')
@login_required
def attendance_timeclock():
    user = User.query.get(session['user_id'])
    
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('under_development.html', user=user, page_title="Time Clock")

@attendance_bp.route('/mobile')
@login_required
def attendance_mobile():
    user = User.query.get(session['user_id'])
    
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('under_development.html', user=user, page_title="Mobile Attendance")

@attendance_bp.route('/weekly')
@login_required
def attendance_weekly():
    user = User.query.get(session['user_id'])
    
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('under_development.html', user=user, page_title="Weekly Attendance")

@attendance_bp.route('/export')
@login_required
def attendance_export():
    user = User.query.get(session['user_id'])
    
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('under_development.html', user=user, page_title="Export Attendance")

@attendance_bp.route('/analytics')
@login_required
def attendance_analytics():
    user = User.query.get(session['user_id'])
    
    if not user:
        session.pop('user_id', None)
        return redirect('/login')
    
    return render_template('under_development.html', user=user, page_title="Attendance Analytics")